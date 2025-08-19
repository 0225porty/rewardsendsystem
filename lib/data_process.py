import __main__
import configparser
import pandas as pd
import csv
import shutil
from tkinter import filedialog
import os
import datetime as dt
from etc import common
import openpyxl
from openpyxl.styles import PatternFill
import inspect

myEnv = common.env_setting[os.path.basename(__main__.__file__)]
mapping = configparser.ConfigParser()
edit = configparser.ConfigParser()
mapping.read(str(myEnv['mapping_path']).replace("{base}",os.path.dirname(__main__.__file__)))
edit.read(str(myEnv['edit_path']).replace("{base}",os.path.dirname(__main__.__file__)))

def create_base_path(params):
    
    # 出力先のフォルダを指定
    myEnv['export_path'] = str(myEnv['export_path']).replace("{campaign_flg}",str(params['campaign_flg']))
    myEnv['src_path'] = str(myEnv['src_path']).replace("{export_path}",myEnv['export_path'])
    myEnv['upload_path'] = str(myEnv['upload_path']).replace("{export_path}",myEnv['export_path'])
    
    # 出力先のベースパスなければ作成
    if not(os.path.isdir(myEnv['export_path'])):
        os.makedirs(myEnv['export_path'])

    # 「src」フォルダなければ作成
    if not(os.path.isdir(myEnv['src_path'])):
        os.makedirs(myEnv['src_path'])

    # 「upload」フォルダなければ作成
    if not(os.path.isdir(myEnv['upload_path'])):
        os.makedirs(myEnv['upload_path'])
    
    # log出力先ファイルを取得
    log_name = common.log_files[params['mode']]
    myEnv['log_path'] = str(myEnv['log_path']).replace("{base}",os.path.dirname(__main__.__file__))
    wb = openpyxl.load_workbook(os.path.join(myEnv['log_path'],str(log_name)))
    
    # log出力先ファイルをファイル名=変換して、出力先ベースパスへコピー
    myEnv['log_name'] = str(log_name).replace(".xlsx",f"_{dt.datetime.now().strftime('%Y%m%d%H%M%S')}.xlsx")
    wb.save(os.path.join(myEnv['export_path'],myEnv['log_name']))
    output_log(params,1,f"logファイル作成　出力先ベースパス：{myEnv['export_path']}　logファイル名：{myEnv['log_name']}")
    output_log(params,1,f"出力先のベースパス作成：{myEnv['export_path']}")
    output_log(params,1,f"「src」パスの作成：{myEnv['src_path']}")
    output_log(params,1,f"「upload」パスの作成：{myEnv['upload_path']}")

def output_log(params,level,msg):

    wb = openpyxl.load_workbook(os.path.join(myEnv['export_path'],str(myEnv['log_name'])))
    ws = wb[str(myEnv['sheet_name'])]
    
    last_row = ws.max_row   # 最終行を取得
    if level == 1:
        ws.cell(row=last_row+1,column=2).value = "情報"
        ws.cell(row=last_row+1,column=3).value = f"{inspect.stack()[1].function}:{inspect.stack()[1].lineno}"
        ws.cell(row=last_row+1,column=4).value = msg
    elif level == 2:
        ws.cell(row=last_row+1,column=2).value = "注意"
        ws.cell(row=last_row+1,column=3).value = f"{inspect.stack()[1].function}:{inspect.stack()[1].lineno}"
        ws.cell(row=last_row+1,column=4).value = msg
    elif level == 3:
        ws.cell(row=last_row+1,column=2).value = "警告"
        ws.cell(row=last_row+1,column=3).value = f"{inspect.stack()[1].function}:{inspect.stack()[1].lineno}"
        ws.cell(row=last_row+1,column=4).value = msg
    
    # ログの保存
    wb.save(os.path.join(myEnv['export_path'],str(myEnv['log_name'])))

def read_folders():
    iDir = os.path.abspath(os.path.dirname(__file__))
    iDirPath = filedialog.askdirectory(initialdir=iDir)
    if iDirPath != '':
        return iDirPath
    else:
        return None

def read_files(params,path):

    output_log(params,1,f"{read_files.__name__}[start]")
    output_log(params,1,f"読み込み先パス：{path}")

    # 指定したフォルダ直下に存在するファイル名をすべて取得
    files = [f for f in os.listdir(path) if os.path.isfile(os.path.join(path,f))]
    output_log(params,1,f"ファイル読み込み完了：{files}")

    # 元ファイルを「src」フォルダへコピー
    for f in files:
        shutil.copy2(os.path.join(path,f),myEnv['src_path'])
        output_log(params,1,f"「src」フォルダへコピー完了：{f}")

    output_log(params,1,f"{read_files.__name__}[end]")
    return files

def set_params(params,df:pd):
    
    output_log(params,1,f"{set_params.__name__}[start]")
    
    if str(params['mode']).split('@')[1] != 'プロモコード':
        col_cnt = df.shape[1]
        new_col = ['column' + str(i) for i in range(col_cnt)]
        df.columns = new_col
        output_log(params,1,f"元データのカラム名変更：{new_col}")

    if params['campaign_id'] != "":
        df['campaign_id'] = params['campaign_id']
        df['campaign_flg'] = params['campaign_flg']
        output_log(params,1,f"キャンペーンID設定完了：{params['campaign_id']}")
    if params['Rewardname_Description'] != "":
        df['Rewardname_Description'] = params['Rewardname_Description']
        output_log(params,1,f"特典説明の設定完了：{params['Rewardname_Description']}")
    if params['useby_date'] != "":
        df['useby_date'] = params['useby_date']
        output_log(params,1,f"特典有効期限の設定完了：{params['useby_date']}")
    
    output_log(params,1,f"{set_params.__name__}[end]")
    
def move_data(params,mapping_rule,df:pd.DataFrame):

    output_log(params,1,f"{move_data.__name__}[start]")
    
    # [mapping.ini]の設定に従って、データを移行
    if(not(mapping.has_section(mapping_rule))):
        output_log(params,2,f"「mapping.ini」にセクション名なし：{mapping_rule}")
        return
    output_log(params,1,f"セクション名：{mapping_rule}")

    src_col=mapping[mapping_rule]['src']
    format=mapping[mapping_rule]['format']
    dtype=mapping[mapping_rule]['dtype']
    output_log(params,1,f"移行元カラム：{src_col}／データ型：{dtype}／データ形式：{format}")

    if(dtype=='str'):
        df_dest = df[src_col].str.strip()
    elif(dtype=='int') or (dtype=='bool'):
        df_dest = df[src_col]
    elif(dtype=='datetime'):
        df_dest = [s.strftime(format) for s in pd.to_datetime(df[src_col])]
    
    output_log(params,1,f"{move_data.__name__}[end]")
    return df_dest

def format_data(params,edit_rule,df:pd.DataFrame):

    output_log(params,1,f"{format_data.__name__}[start]")
    
    # [edit.ini]に該当セクションがあるかを確認
    if edit.has_section(edit_rule):
        
        output_log(params,1,f"セクション名：{edit_rule}")        
        
        colname=edit[edit_rule]['col']
        add=edit[edit_rule]['add']
        substi=pd.DataFrame(data=eval(edit[edit_rule]['substi']))
        output_log(params,1,f"対象カラム名：{colname}／追加文字：{add}／置換文字：{substi}")
    else:
        output_log(params,2,f"「edit.ini」にセクション名なし：{edit_rule}")

    # 特典に回数分の文字を付加する + 「benefit for:」の文字を削除
    output_log(params,2,f"特典名に回数分の文字を付加")
    df['RewardName'] = add_maxlimit(params,df)
    
    output_log(params,1,f"{format_data.__name__}[end]")
    return df

def shuffle_date(params,df:pd):
    
    output_log(params,1,f"{shuffle_date.__name__}[start]")
    df = df.sample(frac=1,ignore_index=True)
    output_log(params,1,f"{shuffle_date.__name__}[end]")
    return df

def export_data(params,df:pd,filename):
    
    output_log(params,1,f"{export_data.__name__}[start]")

    # csv形式で出力する
    # 出力形式は「UTF-8」、ダブルクォートは最小限
    export_file = str(myEnv['export_file']).replace("{filename}",filename)
    df.to_csv(os.path.join(myEnv['upload_path'],export_file),index=False,encoding="utf-8",quoting=0)
    output_log(params,1,f"特典のcsv出力結果　出力先フォルダ：{myEnv['upload_path']}　ファイル名：{export_file}")
    output_log(params,1,f"{export_data.__name__}[end]")

def add_maxlimit(params,df:pd.DataFrame):
    
    update_names = []
    for _,row in df.iterrows():
        update_name = []
        max_limit = pd.to_numeric(row['MaxLimit'])
        if max_limit > 1:
            update_name = str(row['RewardName']).replace("Reward for benefits  :[","").replace("]","") + str(max_limit) + "回分"
        else:
            update_name = str(row['RewardName']).replace("Reward for benefits  :[","").replace("]","")
        update_names.append(update_name)

    return pd.DataFrame(update_names)
